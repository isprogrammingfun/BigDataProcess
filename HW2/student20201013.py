#!/usr/bin/python3
import openpyxl

wb = openpyxl.load_workbook("student.xlsx")
ws = wb['Sheet1']
grade_lists = []
sorted_lists = []

row_id = 1
for row in ws:
    if row_id != 1:
        sum_v = ws.cell(row = row_id, column = 3).value * 0.3
        sum_v += ws.cell(row = row_id, column = 4).value * 0.35
        sum_v += ws.cell(row = row_id, column = 5).value * 0.34
        sum_v += ws.cell(row = row_id, column = 6).value 
        ws.cell(row = row_id, column = 7).value = sum_v
        grade_lists.append([row_id, sum_v, False])
    row_id += 1

sorted_lists = sorted(grade_lists, key = lambda x:x[1], reverse = True)

for i in range(len(sorted_lists)):
    if i <= len(sorted_lists) * 0.15:
            if (sorted_lists[i-1][1] == sorted_lists[i][1]) and (sorted_lists[i+1][1] == sorted_lists[i][1]):
                sorted_lists[i][2] = True
            ws.cell(row = sorted_lists[i][0], column = 8).value = 'A+'

    elif i <= len(sorted_lists) * 0.3:
        if (sorted_lists[i-1][1] == sorted_lists[i][1]) and (sorted_lists[i+1][1] == sorted_lists[i][1]):
                sorted_lists[i][2] = True
        ws.cell(row = sorted_lists[i][0], column = 8).value = "A0"

    elif i <= len(sorted_lists) * 0.5:
        if (sorted_lists[i-1][1] == sorted_lists[i][1]) and (sorted_lists[i+1][1] == sorted_lists[i][1]):
                sorted_lists[i][2] = True
        ws.cell(row = sorted_lists[i][0], column = 8).value = "B+"
    
    elif i <= len(sorted_lists) * 0.7:
        if (sorted_lists[i-1][1] == sorted_lists[i][1]) and (sorted_lists[i+1][1] == sorted_lists[i][1]):
                sorted_lists[i][2] = True
        ws.cell(row = sorted_lists[i][0], column = 8).value = "B0"

    elif i <= len(sorted_lists) * 0.85:
        if (sorted_lists[i-1][1] == sorted_lists[i][1]) and (sorted_lists[i+1][1] == sorted_lists[i][1]):
                sorted_lists[i][2] = True
        ws.cell(row = sorted_lists[i][0], column = 8).value = "C+"

    elif i <= len(sorted_lists) * 1:
        if (sorted_lists[i-1][1] == sorted_lists[i][1]) and (sorted_lists[i+1][1] == sorted_lists[i][1]):
                sorted_lists[i][2] = True
        ws.cell(row = sorted_lists[i][0], column = 8).value = "C0"
i += 1
                              
wb.save("student.xlsx")
wb.close()